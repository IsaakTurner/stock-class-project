import os
import argparse
import logging
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import requests
import requests_cache
from scipy.stats import linregress
import datetime

# ---------------------- Setup Logging & Caching ----------------------
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
# Cache API requests for 5 minutes
requests_cache.install_cache("alpha_vantage_cache", expire_after=300)

# ---------------------- Optional Libraries ----------------------
try:
    import mplfinance as mpf
except ImportError:
    logging.warning("mplfinance not installed. Candlestick charts will not be available.")
    mpf = None

try:
    from prophet import Prophet
except ImportError:
    logging.warning("Prophet not installed. Forecasting will be done via linear regression.")
    Prophet = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as ExcelImage
except ImportError:
    logging.error("openpyxl is required for Excel output. Please install it.")
    exit(1)

# ---------------------- Alpha Vantage Library ----------------------
try:
    from alpha_vantage.timeseries import TimeSeries
except ImportError:
    logging.error("alpha_vantage module not installed. Please install via pip install alpha_vantage")
    exit(1)

# ---------------------- Data Fetching Functions ----------------------
def fetch_stock_data(symbol, api_key, start_date=None, end_date=None):
    """Fetch stock data using Alpha Vantage API; use cache if available."""
    cache_dir = "cache"
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
    cache_file = os.path.join(cache_dir, f"{symbol}_data.csv")
    
    # Try to load cached data
    if os.path.exists(cache_file):
        try:
            stock_data = pd.read_csv(cache_file, parse_dates=["date"], index_col="date")
            logging.info(f"Loaded cached data for {symbol}")
        except Exception as e:
            logging.warning(f"Error reading cache file: {e}. Fetching new data.")
            stock_data = None
    else:
        stock_data = None

    if stock_data is None:
        ts = TimeSeries(key=api_key, output_format="pandas")
        try:
            data, meta = ts.get_daily(symbol=symbol, outputsize="full")
            data = data.rename(columns={
                "1. open": "Open",
                "2. high": "High",
                "3. low": "Low",
                "4. close": "Close",
                "5. volume": "Volume"
            })
            # Reverse order (oldest first) and convert index to datetime
            data = data[::-1]
            data.index = pd.to_datetime(data.index)
            data.ffill(inplace=True)
            stock_data = data
            stock_data.to_csv(cache_file, index_label="date")
            logging.info(f"Fetched and cached new data for {symbol}")
        except Exception as e:
            logging.error(f"Error fetching data for {symbol}: {e}")
            return None

    # Filter by date if provided
    if start_date:
        stock_data = stock_data[stock_data.index >= pd.to_datetime(start_date)]
    if end_date:
        stock_data = stock_data[stock_data.index <= pd.to_datetime(end_date)]
    
    return stock_data

def fetch_index_data(symbol, api_key, start_date=None, end_date=None):
    """Fetch index data (e.g., SPY) as a proxy for the S&P 500."""
    ts = TimeSeries(key=api_key, output_format="pandas")
    try:
        data, meta = ts.get_daily(symbol=symbol, outputsize="full")
        data = data.rename(columns={
            "1. open": "Open",
            "2. high": "High",
            "3. low": "Low",
            "4. close": "Close",
            "5. volume": "Volume"
        })
        data = data[::-1]
        data.index = pd.to_datetime(data.index)
        data.ffill(inplace=True)
        if start_date:
            data = data[data.index >= pd.to_datetime(start_date)]
        if end_date:
            data = data[data.index <= pd.to_datetime(end_date)]
        logging.info(f"Fetched index data for {symbol}")
        return data
    except Exception as e:
        logging.error(f"Error fetching index data for {symbol}: {e}")
        return None

def fetch_earnings_data(symbol, api_key):
    """Fetch quarterly earnings data using Alpha Vantage API."""
    url = f"https://www.alphavantage.co/query?function=EARNINGS&symbol={symbol}&apikey={api_key}"
    try:
        response = requests.get(url)
        earnings_data = response.json()
        earnings_dates = []
        if "quarterlyEarnings" in earnings_data:
            earnings_dates = [
                pd.to_datetime(q["reportedDate"]) for q in earnings_data["quarterlyEarnings"] if "reportedDate" in q
            ]
            logging.info(f"Fetched earnings dates for {symbol}: {earnings_dates[:4]}")
        else:
            logging.warning("No quarterly earnings data found.")
        return earnings_data, earnings_dates
    except Exception as e:
        logging.warning(f"Error fetching earnings data: {e}")
        return None, []

# ---------------------- Technical Indicator Functions ----------------------

def compute_Bollinger_Bands(stock_data, window=20, num_std=2):
    """Compute Bollinger Bands for the stock data."""
    stock_data['BB_Mid'] = stock_data['Close'].rolling(window=window).mean()
    stock_data['BB_Upper'] = stock_data['BB_Mid'] + (num_std * stock_data['Close'].rolling(window=window).std())
    stock_data['BB_Lower'] = stock_data['BB_Mid'] - (num_std * stock_data['Close'].rolling(window=window).std())
    return stock_data

def compute_MACD(stock_data, short_window=12, long_window=26, signal_window=9):
    """Compute MACD (Moving Average Convergence Divergence) and Signal Line."""
    stock_data['MACD'] = stock_data['Close'].ewm(span=short_window, adjust=False).mean() - \
                          stock_data['Close'].ewm(span=long_window, adjust=False).mean()
    stock_data['MACD_Signal'] = stock_data['MACD'].ewm(span=signal_window, adjust=False).mean()
    return stock_data

def compute_RSI(stock_data, window=14):
    """Compute Relative Strength Index (RSI) for a given stock."""
    delta = stock_data['Close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
    rs = gain / loss
    stock_data['RSI'] = 100 - (100 / (1 + rs))
    return stock_data

def compute_moving_averages(stock_data):
    """Compute moving averages for the stock data."""
    stock_data['50_MA'] = stock_data['Close'].rolling(window=50).mean()
    stock_data['100_MA'] = stock_data['Close'].rolling(window=100).mean()
    stock_data['200_MA'] = stock_data['Close'].rolling(window=200).mean()
    return stock_data

def compute_technical_indicators(stock_data):
    """Compute all technical indicators for the stock data."""
    stock_data = compute_moving_averages(stock_data)
    stock_data = compute_RSI(stock_data)
    stock_data = compute_Bollinger_Bands(stock_data)
    stock_data = compute_MACD(stock_data)
    return stock_data  
# ---------------------- Charting Functions ----------------------
def plot_price_chart(stock_data, stock_symbol, earnings_dates, filename):
    """Plot a line chart of the stock price with moving averages and earnings dates."""
    plt.figure(figsize=(14, 7))
    plt.plot(stock_data.index, stock_data["Close"], label="Close Price", color="black")
    plt.plot(stock_data.index, stock_data["50_MA"], label="50-Day MA", color="blue", linestyle="--")
    plt.plot(stock_data.index, stock_data["100_MA"], label="100-Day MA", color="green", linestyle="--")
    plt.plot(stock_data.index, stock_data["200_MA"], label="200-Day MA", color="red", linestyle="--")
    for ed in earnings_dates:
        if ed in stock_data.index:
            plt.axvline(x=ed, color="purple", linestyle=":", alpha=0.7)
    plt.title(f"{stock_symbol} Price Chart with Moving Averages")
    plt.xlabel("Date")
    plt.ylabel("Price (USD)")
    plt.legend()
    plt.grid()
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()
    logging.info(f"Saved price chart as {filename}")

def plot_candlestick_chart(stock_data, stock_symbol, filename):
    """Plot a candlestick chart with an RSI subplot using mplfinance."""
    if mpf is None:
        logging.warning("mplfinance not available. Skipping candlestick chart.")
        return
    try:
        data_for_candle = stock_data[-500:]
        mc = mpf.make_marketcolors(up="g", down="r", inherit=True)
        style = mpf.make_mpf_style(marketcolors=mc)
        add_plots = [mpf.make_addplot(data_for_candle["RSI"], panel=1, color="blue", ylabel="RSI")]
        mpf.plot(
            data_for_candle,
            type="candle",
            style=style,
            addplot=add_plots,
            title=f"{stock_symbol} Candlestick Chart",
            volume=True,
            warn_too_much_data=1000,
            savefig=filename,
        )
        logging.info(f"Saved candlestick chart as {filename}")
    except Exception as e:
        logging.error(f"Error generating candlestick chart: {e}")

# ---------------------- Forecasting Functions ----------------------
def simple_linear_forecast(stock_data, forecast_days):
    """Simple linear regression forecasting as a fallback."""
    slope, intercept, _, _, _ = linregress(range(len(stock_data)), stock_data["Close"])
    future_dates = [stock_data.index[-1] + pd.Timedelta(days=i) for i in range(1, forecast_days + 1)]
    future_prices = [slope * (len(stock_data) + i) + intercept for i in range(1, forecast_days + 1)]
    forecast_df = pd.DataFrame({"ds": future_dates, "yhat": future_prices})
    return forecast_df

def forecast_prices(stock_data, forecast_days, stock_symbol, filename):
    """
    Forecast future prices using Prophet if available.
    Falls back to simple linear regression if Prophet is not available or fails.
    """
    if Prophet is not None:
        try:
            forecast_df = stock_data.reset_index().rename(columns={"date": "ds", "Close": "y"})
            m = Prophet(daily_seasonality=True)
            m.fit(forecast_df[['ds', 'y']])
            future = m.make_future_dataframe(periods=forecast_days)
            forecast = m.predict(future)
            fig = m.plot(forecast)
            fig.savefig(filename)
            plt.close(fig)
            logging.info("Forecasting completed using Prophet.")
            return forecast
        except Exception as e:
            logging.error(f"Error during forecasting with Prophet: {e}")
    logging.warning("Using simple linear regression for forecasting.")
    forecast = simple_linear_forecast(stock_data, forecast_days)
    # Plot the simple forecast
    plt.figure(figsize=(10, 6))
    plt.plot(stock_data.index, stock_data["Close"], label="Historical Price")
    plt.plot(forecast["ds"], forecast["yhat"], label="Forecast", color="orange")
    plt.title(f"{stock_symbol} Forecast (Linear Regression)")
    plt.xlabel("Date")
    plt.ylabel("Price")
    plt.legend()
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()
    return forecast
#----------------------- Monte Carlo Simmulation -------------------
'''def monte_carlo_simulation(stock_data, symbol, num_simulations=1000, forecast_days=252, save_path=None):
    """
    Run a Monte Carlo simulation for stock price prediction using Geometric Brownian Motion (GBM).
    
    Parameters:
        stock_data (pd.DataFrame): DataFrame containing historical stock prices.
        symbol (str): Stock symbol.
        num_simulations (int): Number of Monte Carlo simulations to run.
        forecast_days (int): Number of days to forecast.
        save_path (str, optional): Path to save the generated plot.
    """
    if stock_data is None or stock_data.empty:
        print("Error: No stock data available for Monte Carlo simulation.")
        return None, None
    
    last_price = stock_data['Close'].iloc[-1]
    log_returns = np.log(stock_data['Close'] / stock_data['Close'].shift(1)).dropna()
    mu = log_returns.mean()
    sigma = log_returns.std()
    
    simulations = np.zeros((forecast_days, num_simulations))
    for i in range(num_simulations):
        price_series = [last_price]
        for _ in range(forecast_days - 1):
            price_series.append(price_series[-1] * np.exp(np.random.normal(mu - (sigma**2) / 2, sigma)))
        simulations[:, i] = price_series
    
    plt.figure(figsize=(12, 6))
    plt.plot(simulations, color='gray', alpha=0.1)
    plt.plot(np.mean(simulations, axis=1), color='red', label='Mean Projection')
    plt.axhline(y=last_price, color='black', linestyle='--', label='Last Close Price')
    plt.xlabel('Days')
    plt.ylabel('Price (USD)')
    plt.title(f'Monte Carlo Simulation: {symbol}')
    plt.legend()
    plt.grid()
    
    if save_path:
        plt.savefig(save_path)
        print(f"Monte Carlo simulation chart saved as {save_path}")
    
    plt.show()
    
    # Confidence intervals
    percentiles = np.percentile(simulations[-1, :], [5, 50, 95])
    print(f"Monte Carlo Forecast for {symbol} in {forecast_days} days:")
    print(f"  - 5th Percentile (Bearish Case): ${percentiles[0]:.2f}")
    print(f"  - Median Price (Most Likely): ${percentiles[1]:.2f}")
    print(f"  - 95th Percentile (Bullish Case): ${percentiles[2]:.2f}")
    
    return simulations, percentiles'''
# ---------------------- Backtesting Strategy ----------------------
def backtest_strategy(stock_data, initial_balance=10000):
    """Backtest a simple trading strategy using RSI."""
    balance = initial_balance
    position = 0
    for i in range(1, len(stock_data)):
        if stock_data['RSI'].iloc[i] < 30 and balance > 0:
            # Buy signal (RSI < 30)
            position = balance / stock_data['Close'].iloc[i]
            balance = 0
        elif stock_data['RSI'].iloc[i] > 70 and position > 0:
            # Sell signal (RSI > 70)
            balance = position * stock_data['Close'].iloc[i]
            position = 0
    final_value = balance + (position * stock_data['Close'].iloc[-1])
    return final_value, (final_value - initial_balance) / initial_balance * 100
# ---------------------- Performance & Valuation Functions ----------------------
def calculate_performance_metrics(stock_data, index_data):
    """Calculate additional performance metrics (Sharpe ratio, alpha, beta)."""
    stock_data['daily_return'] = stock_data["Close"].pct_change()
    
    if index_data is not None:
        index_data['daily_return'] = index_data["Close"].pct_change()

        # Align both datasets to have matching dates
        stock_returns, index_returns = stock_data['daily_return'].align(index_data['daily_return'], join='inner')

        print(f"Length of stock_returns: {len(stock_returns)}")
        print(f"Length of index_returns: {len(index_returns)}")

        # Handle missing values explicitly
        valid_data = pd.concat([stock_returns, index_returns], axis=1).dropna()
        stock_returns = valid_data.iloc[:, 0]
        index_returns = valid_data.iloc[:, 1]

        print(f"Length after dropping NaN: {len(stock_returns)}, {len(index_returns)}")

        # Assume a risk-free rate of 0.02/252 (daily)
        risk_free_rate = 0.02 / 252
        excess_return = stock_returns - risk_free_rate
        sharpe_ratio = excess_return.mean() / excess_return.std() if excess_return.std() != 0 else np.nan

        if len(stock_returns) == len(index_returns) and len(stock_returns) > 1:
            beta, alpha, r_value, p_value, std_err = linregress(index_returns, stock_returns)
            return sharpe_ratio, alpha, beta
        else:
            print("Error: Stock and index returns have mismatched lengths or insufficient data.")
            return None, None, None
    return None, None, None

def analyze_valuation(stock_data, earnings_data, target_pe, index_data):
    """Perform EPS valuation and performance analysis."""
    valuation_analysis = "Valuation Analysis:\n"
    
    # EPS Valuation
    if earnings_data and "quarterlyEarnings" in earnings_data:
        try:
            earnings_df = pd.DataFrame(earnings_data["quarterlyEarnings"])
            if "reportedEPS" in earnings_df.columns:
                earnings_df = earnings_df[earnings_df["reportedEPS"].notnull()]
                earnings_df["reportedEPS"] = earnings_df["reportedEPS"].astype(float)
                earnings_df["reportedDate"] = pd.to_datetime(earnings_df["reportedDate"])
                earnings_df = earnings_df.sort_values("reportedDate", ascending=False)
                latest_eps = earnings_df.iloc[0]["reportedEPS"]
                valuation_analysis += f"Latest Quarterly EPS: {latest_eps:.2f}\n"
                recent_eps = earnings_df.head(4)["reportedEPS"]
                avg_quarterly_eps = recent_eps.mean()
                annualized_eps = avg_quarterly_eps * 4
                valuation_analysis += f"Annualized EPS (avg of last 4 quarters * 4): {annualized_eps:.2f}\n"
                valuation_analysis += f"Assumed Target P/E Ratio: {target_pe}\n"
                fair_value = annualized_eps * target_pe
                valuation_analysis += f"Estimated Fair Value: {fair_value:.2f}\n"
                current_price = stock_data["Close"].iloc[-1]
                valuation_analysis += f"Current Stock Price: {current_price:.2f}\n"
                diff = fair_value - current_price
                valuation_analysis += f"Difference (Fair Value - Current Price): {diff:.2f}\n"
                if diff > 0:
                    valuation_analysis += "Stock appears undervalued.\n"
                else:
                    valuation_analysis += "Stock appears overvalued.\n"
            else:
                valuation_analysis += "EPS data not available for valuation.\n"
        except Exception as e:
            logging.warning(f"Error computing EPS valuation: {e}")
            valuation_analysis += "EPS data unavailable for valuation.\n"
    else:
        valuation_analysis += "No quarterly earnings data available for valuation.\n"
    
    # Performance Comparison with S&P 500 (using SPY)
    if index_data is not None:
        common_dates = stock_data.index.intersection(index_data.index)
        if not common_dates.empty:
            stock_common = stock_data.loc[common_dates]
            index_common = index_data.loc[common_dates]
            stock_return = (stock_common["Close"].iloc[-1] / stock_common["Close"].iloc[0]) - 1
            index_return = (index_common["Close"].iloc[-1] / index_common["Close"].iloc[0]) - 1
            valuation_analysis += "\nPerformance Comparison (Common Period):\n"
            valuation_analysis += f"Stock Total Return: {stock_return*100:.2f}%\n"
            valuation_analysis += f"S&P 500 Total Return: {index_return*100:.2f}%\n"
            if stock_return > index_return:
                valuation_analysis += "The stock outperformed the S&P 500.\n"
            else:
                valuation_analysis += "The stock underperformed the S&P 500.\n"
            
            sharpe, alpha, beta = calculate_performance_metrics(stock_data, index_data)
            if sharpe is not None:
                valuation_analysis += f"\nSharpe Ratio: {sharpe:.2f}\nAlpha: {alpha:.4f}\nBeta: {beta:.4f}\n"
        else:
            valuation_analysis += "\nNo overlapping dates for performance comparison.\n"
    else:
        valuation_analysis += "\nS&P 500 data not available for performance comparison.\n"
    
    return valuation_analysis

# ---------------------- Excel Output Function ----------------------
def save_to_excel(stock_data, valuation_analysis, chart_files, stock_symbol, output_format="xlsx"):
    """Save stock data, valuation analysis, and charts to an Excel file (or CSV)."""
    if output_format == "csv":
        csv_file = f"{stock_symbol}_stock_data.csv"
        stock_data_reset = stock_data.copy().reset_index()
        stock_data_reset.to_csv(csv_file, index=False)
        logging.info(f"Data saved as CSV: {csv_file}")
        return csv_file
    else:
        excel_file = f"{stock_symbol}_stock_data.xlsx"
        # Remove existing file if present
        if os.path.exists(excel_file):
            os.remove(excel_file)
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            stock_data_reset = stock_data.copy().reset_index()
            if stock_data_reset.columns[0] != "Date":
                stock_data_reset.rename(columns={stock_data_reset.columns[0]: "Date"}, inplace=True)
            stock_data_reset.sort_values("Date", ascending=False, inplace=True)
            stock_data_reset.to_excel(writer, sheet_name="Stock Data", index=False)
            valuation_df = pd.DataFrame({"Valuation Analysis": valuation_analysis.split("\n")})
            valuation_df.to_excel(writer, sheet_name="Valuation", index=False)
        
        # Load workbook for further formatting and chart insertion
        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception as e:
            logging.error(f"Error loading the workbook: {e}")
            return excel_file
        
        # ---- Format STOCK DATA Sheet ----
        ws_stock = wb["Stock Data"]
        ws_stock.insert_rows(1)
        ws_stock["A1"] = f"{stock_symbol} Stock Data"
        ws_stock["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws_stock["A1"].alignment = Alignment(horizontal="center")
        ws_stock.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ws_stock[2]))
        title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws_stock["A1"].fill = title_fill

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for cell in ws_stock[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        for col in ws_stock.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            col_letter = get_column_letter(col[0].column)
            ws_stock.column_dimensions[col_letter].width = max_length + 3

        ws_stock.freeze_panes = "A3"
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in ws_stock.iter_rows(min_row=3, max_row=ws_stock.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill

        # ---- Format VALUATION Sheet ----
        ws_val = wb["Valuation"]
        ws_val["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws_val["A1"].alignment = Alignment(horizontal="center")
        ws_val["A1"].fill = title_fill
        for col in ws_val.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            col_letter = get_column_letter(col[0].column)
            ws_val.column_dimensions[col_letter].width = max_length + 5

        # ---- Insert Chart Images into a New Sheet ----
        ws_charts = wb.create_sheet("Charts")
        if "price_chart" in chart_files and os.path.exists(chart_files["price_chart"]):
            img_price = ExcelImage(chart_files["price_chart"])
            img_price.width, img_price.height = 640, 480
            ws_charts.add_image(img_price, "A1")
        if "candlestick_chart" in chart_files and os.path.exists(chart_files["candlestick_chart"]):
            img_candle = ExcelImage(chart_files["candlestick_chart"])
            img_candle.width, img_candle.height = 640, 480
            ws_charts.add_image(img_candle, "A25")
        if "forecast_chart" in chart_files and os.path.exists(chart_files["forecast_chart"]):
            img_forecast = ExcelImage(chart_files["forecast_chart"])
            img_forecast.width, img_forecast.height = 640, 480
            ws_charts.add_image(img_forecast, "A49")

        wb.save(excel_file)
        logging.info(f"Data, charts, and valuation saved to Excel file: {excel_file}")
        return excel_file

# ---------------------- Main Script ----------------------
def main():
    parser = argparse.ArgumentParser(description="Enhanced Stock Data Analysis Script")
    parser.add_argument("--api_key", type=str, default=os.environ.get("ALPHA_VANTAGE_API_KEY", "R1ECHASVMAA0UE87"),
                        help="Alpha Vantage API Key")
    parser.add_argument("--symbol", type=str, default="NKE", help="Stock ticker symbol (e.g., AAPL, TSLA)")
    parser.add_argument("--target_pe", type=float, default=15.0, help="Target P/E ratio for valuation")
    parser.add_argument("--start_date", type=str, help="Start date for analysis (YYYY-MM-DD)")
    parser.add_argument("--end_date", type=str, help="End date for analysis (YYYY-MM-DD)")
    parser.add_argument("--chart_type", type=str, choices=["line", "candlestick"], default="line",
                        help="Type of stock chart")
    parser.add_argument("--forecast_days", type=int, default=30,
                        help="Days to forecast using Prophet or linear regression")
    parser.add_argument("--output_format", type=str, choices=["xlsx", "csv"], default="xlsx",
                        help="Output file format")
    args = parser.parse_args()
    stock_data = fetch_stock_data(args.symbol, args.api_key, args.start_date, args.end_date)
    stock_data = compute_technical_indicators(stock_data)
    chart_file = f"{args.symbol}_chart_with_indicators.png"
    plot_price_chart(stock_data, args.symbol, [], chart_file)
    final_balance, return_percentage = backtest_strategy(stock_data)
    logging.info(f"Backtesting Results: Final Balance: ${final_balance:.2f}, Return: {return_percentage:.2f}%")

    # Fetch stock data
    stock_data = fetch_stock_data(args.symbol, args.api_key, args.start_date, args.end_date)
    if stock_data is None or stock_data.empty:
        logging.error("No stock data available. Exiting.")
        exit(1)
    
    # Fetch index data (using SPY as a proxy for S&P 500)
    index_data = fetch_index_data("SPY", args.api_key, args.start_date, args.end_date)
    
    # Fetch earnings data
    earnings_data, earnings_dates = fetch_earnings_data(args.symbol, args.api_key)
    
    # Compute technical indicators (moving averages, RSI)
    stock_data = compute_technical_indicators(stock_data)
    
    # Generate charts
    chart_files = {}
    price_chart_file = f"{args.symbol}_price_chart.png"
    plot_price_chart(stock_data, args.symbol, earnings_dates, price_chart_file)
    chart_files["price_chart"] = price_chart_file
    #simulations, percentiles = monte_carlo_simulation(stock_data, symbol="AAPL", save_path="monte_carlo_aapl.png")
    
    if args.chart_type == "candlestick":
        candlestick_chart_file = f"{args.symbol}_candlestick_chart.png"
        plot_candlestick_chart(stock_data, args.symbol, candlestick_chart_file)
        chart_files["candlestick_chart"] = candlestick_chart_file

    forecast_chart_file = f"{args.symbol}_forecast_chart.png"
    forecast = forecast_prices(stock_data, args.forecast_days, args.symbol, forecast_chart_file)
    chart_files["forecast_chart"] = forecast_chart_file
    
    # Analyze valuation and performance
    valuation_analysis = analyze_valuation(stock_data, earnings_data, args.target_pe, index_data)
    
    # Save results to Excel (or CSV)
    output_file = save_to_excel(stock_data, valuation_analysis, chart_files, args.symbol, args.output_format)
    logging.info(f"Analysis complete. Output saved to: {output_file}")

if __name__ == "__main__":
    main()